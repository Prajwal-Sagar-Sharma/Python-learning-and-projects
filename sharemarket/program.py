import bs4
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests
from openpyxl.styles import PatternFill 

class ShareMarket:
    def __init__(self):
        self.url="https://merolagani.com/LatestMarket.aspx"
        self.website=requests.get(self.url)
        self.website.raise_for_status()
        self.website_text=self.website.text
        self.workbook=Workbook()
        self.worksheet=self.workbook.active
        self.worksheet.title="Share"

    def scrap_header(self):
        soup=bs4.BeautifulSoup(self.website_text,'html.parser')
        data_container=soup.find(id="ctl00_ContentPlaceHolder1_LiveTrading")
        data_header=data_container.table.thead.tr.contents
        header=[i.text for i in data_header][:7]
        self.worksheet.append(header)


    def excel_writer(self):
        self.scrap_header()
        data=self.scrap_data()
        staring_data=2
        

        for i in data:
            self.worksheet.append(i)
            profit= PatternFill(start_color='00FF00', end_color='00FF00', fill_type="solid")
            loss=PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
            if  float(i[2])>0:
                for a in range(0,7):
                    self.worksheet[staring_data][a].fill=profit
            else:
                for a in range(0,7):
                    self.worksheet[staring_data][a].fill=loss
            staring_data+=1
        self.workbook.save("share.xlsx")



    def scrap_data(self):
        soup=bs4.BeautifulSoup(self.website_text,'html.parser')
        data_container=soup.find(id="ctl00_ContentPlaceHolder1_LiveTrading")
        data_table=data_container.table.tbody.find_all("tr")
        share_data=[[value.text for value in data.contents][:7] for data in data_table]
        return share_data
        
if __name__=="__main__":
    share=ShareMarket()
    share.excel_writer()